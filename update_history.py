"""
Run this script whenever you update Z47 Index.xlsx to regenerate z47_history.csv.

Usage:
    python update_history.py
    python update_history.py --excel "C:\\path\\to\\Z47 Index.xlsx"
"""

import argparse
import csv
import shutil
import tempfile
import os
from datetime import date
import openpyxl

DEFAULT_EXCEL = (
    r"C:\Users\Girish Shenoy\Z47\The Vault - Documents\Corp Dev"
    r"\8. Public market projects and research\IPO CoE\Z47 Index.xlsx"
)
OUTPUT_CSV = os.path.join(os.path.dirname(__file__), "z47_history.csv")


def extract(excel_path: str) -> list[dict]:
    # Copy to temp (in case Excel has the file locked)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    shutil.copy2(excel_path, tmp.name)

    wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
    ws = wb["Z47 Index"]

    rows = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=8, values_only=True):
        dt, z47_float, z47_mcap, nifty, sensex, nifty_abs, sensex_abs = row
        if dt is None or z47_float is None:
            continue
        if hasattr(dt, "date"):
            dt = dt.date()
        rows.append({
            "date":            str(dt),
            "z47_float":       round(z47_float, 4),
            "z47_mcap":        round(z47_mcap, 4)   if z47_mcap   else None,
            "nifty_indexed":   round(nifty, 4)       if nifty      else None,
            "sensex_indexed":  round(sensex, 4)      if sensex     else None,
            "nifty_abs":       round(nifty_abs, 2)   if nifty_abs  else None,
            "sensex_abs":      round(sensex_abs, 2)  if sensex_abs else None,
        })

    wb.close()
    os.unlink(tmp.name)
    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="Path to Z47 Index.xlsx")
    args = parser.parse_args()

    print(f"Reading: {args.excel}")
    rows = extract(args.excel)
    print(f"Extracted {len(rows)} rows  ({rows[0]['date']} → {rows[-1]['date']})")

    with open(OUTPUT_CSV, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    print(f"Saved: {OUTPUT_CSV}")
    print("Restart the Streamlit app (or it will auto-reload within 60 s).")


if __name__ == "__main__":
    main()
